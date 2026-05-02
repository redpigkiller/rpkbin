"""Normalization layer: Excel workbook → InternalGrid.

Responsibilities
----------------
1. Read the workbook via openpyxl (.xlsx/.xlsm) or xlrd (.xls).
   (NOT pandas — pandas cannot detect merge cells.)
2. Build a merge map so that every cell in a merged region carries the master
   cell's value and is flagged `is_merged=True`.
3. Normalise dates/times to ISO-8601 strings (YYYY-MM-DD / HH:MM).
4. Convert coordinates to 0-based for the rest of the engine.
"""

from __future__ import annotations
from dataclasses import dataclass
from typing import Any
import datetime

import xlrd
from xlrd.book import Book
from openpyxl import Workbook


@dataclass
class InternalCell:
    """A single normalised cell.

    value          : str representation of the cell value ("" if empty).
                     Dates are already converted to 'YYYY-MM-DD'.
    original_value : the raw value from the workbook before normalisation.
    is_merged      : True if this cell was expanded from a merge range.
    """
    value: str
    original_value: Any
    is_merged: bool = False


class InternalGrid:
    """A 2-D rectangular array of InternalCell objects with 0-based coordinate access."""

    def __init__(self, cells: list[list[InternalCell]]):
        self._cells = cells
        self.num_rows = len(cells)
        if self.num_rows == 0:
            self.num_cols = 0
            return
        col_widths = set(len(row) for row in cells)
        if len(col_widths) != 1:
            raise ValueError(
                f"InternalGrid requires all rows to have equal length, "
                f"got widths: {sorted(col_widths)}"
            )
        self.num_cols = col_widths.pop()

    def get_cell(self, row: int, col: int) -> InternalCell:
        """Return the cell at (row, col), or None if out of bounds."""
        if row < 0 or row >= self.num_rows:
            raise IndexError("Row index out of range")
        if col < 0 or col >= len(self._cells[row]):
            raise IndexError("Column index out of range")
        return self._cells[row][col]

    def __getitem__(self, index):
        if isinstance(index, tuple):
            if len(index) != 2:
                raise IndexError("Invalid number of indices. Use grid[row, col]")
            
            row_idx, col_idx = index
            
            if isinstance(row_idx, slice):
                rows = self._cells[row_idx]
                if isinstance(col_idx, slice):
                    return [[row[c] for c in range(*col_idx.indices(len(row)))] for row in rows]
                return [row[col_idx] for row in rows]
            
            row = self._cells[row_idx]
            return row[col_idx]

        return self._cells[index]

    def transpose(self) -> "InternalGrid":
        """Return a new InternalGrid with rows and columns swapped."""
        if self.num_rows == 0 or self.num_cols == 0:
            return InternalGrid([])
        transposed: list[list[InternalCell]] = []
        for c in range(self.num_cols):
            row = []
            for r in range(self.num_rows):
                row.append(self._cells[r][c])
            transposed.append(row)
        return InternalGrid(transposed)

# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------
def normalize_value(value):
    """
    Normalize raw excel value to string or None.
    """

    if value is None:
        return ""

    # datetime / date
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")

    # time
    if isinstance(value, datetime.time):
        return value.strftime("%H:%M")

    # float → int
    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    # string
    if isinstance(value, str):
        v = value.strip()
        return v if v else ""

    return str(value)

def _load_xlsx_from_wb(
    wb: Workbook,
    sheet: str,
) -> InternalGrid:
    ws = wb[sheet]

    # Build merge map: (1-based row, 1-based col) → master value
    merge_map = {}

    for crange in ws.merged_cells.ranges:
        rlo = crange.min_row
        rhi = crange.max_row
        clo = crange.min_col
        chi = crange.max_col

        # master is a top-left cell
        master_coord = (rlo, clo)
        for rowx in range(rlo, rhi + 1):
            for colx in range(clo, chi + 1):
                merge_map[(rowx, colx)] = master_coord

    # Build the grid (convert to 0-based)
    max_row = ws.max_row
    max_col = ws.max_column

    grid_cells: list[list[InternalCell]] = []
    for r1 in range(1, max_row + 1):
        row_data: list[InternalCell] = []
        for c1 in range(1, max_col + 1):
            raw_cell = ws.cell(r1, c1)
            is_merged_cell = (r1, c1) in merge_map
            
            if is_merged_cell:
                mr, mc = merge_map[(r1, c1)]
                raw_cell = ws.cell(mr, mc)
            else:
                raw_cell = ws.cell(r1, c1)

            # Post-processing
            norm_val = normalize_value(raw_cell.value)
            internal = InternalCell(value=norm_val, original_value=raw_cell.value, is_merged=is_merged_cell)

            row_data.append(internal)
        grid_cells.append(row_data)

    return InternalGrid(grid_cells)

def _load_xls_from_wb(
    wb: Book,
    sheet: str,
) -> InternalGrid:
    sh = wb.sheet_by_name(sheet)

    # Build merge map: (row, col) 0-based -> master (row, col)
    # xlrd returns merged_cells as list of (row_low, row_high, col_low, col_high)
    merge_map = {}

    for crange in sh.merged_cells:
        rlo, rhi, clo, chi = crange
        # master is a top-left cell
        master_coord = (rlo, clo)
        for rowx in range(rlo, rhi):
            for colx in range(clo, chi):
                merge_map[(rowx, colx)] = master_coord

    max_row = sh.nrows
    max_col = sh.ncols

    grid_cells: list[list[InternalCell]] = []
    for r in range(max_row):
        row_data: list[InternalCell] = []
        for c in range(max_col):
            is_merged_cell = (r, c) in merge_map
            
            if is_merged_cell:
                mr, mc = merge_map[(r, c)]
                raw_val = sh.cell_value(rowx=mr, colx=mc)
                raw_type = sh.cell_type(rowx=mr, colx=mc)
            else:
                raw_val = sh.cell_value(rowx=r, colx=c)
                raw_type = sh.cell_type(rowx=r, colx=c)

            # Post-processing
            if raw_type == xlrd.XL_CELL_DATE and raw_val:
                raw_val = datetime.datetime(
                    *xlrd.xldate_as_tuple(raw_val, wb.datemode)
                )
            norm_val = normalize_value(raw_val)

            internal = InternalCell(value=norm_val, original_value=raw_val, is_merged=is_merged_cell)
            row_data.append(internal)
        grid_cells.append(row_data)

    return InternalGrid(grid_cells)
